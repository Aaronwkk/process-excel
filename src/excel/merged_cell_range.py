import openpyxl
import re
import os
from openpyxl.utils import get_column_letter, column_index_from_string
from dotenv import load_dotenv # 导入 load_dotenv

def unmerge_and_fill_with_original_format(input_filepath: str, output_filepath: str):
    """
    取消合并 Excel 文件中的单元格，并向下填充值，
    同时尝试保留原始单元格的数字格式，特别是百分比列。

    Args:
        input_filepath (str): 包含合并单元格的输入 .xlsx 文件路径。
        output_filepath (str): 保存处理后的 .xlsx 文件的路径。
    """
    try:
        # 加载工作簿
        workbook = openpyxl.load_workbook(input_filepath)
        sheet = workbook.active

        # 获取合并单元格范围的字符串列表
        merged_ranges_str = [str(mr) for mr in sheet.merged_cells]

        # 遍历每个合并单元格范围字符串
        for merged_range_str in merged_ranges_str:
            # 解析单元格范围字符串以获取 min_col, min_row, max_col, max_row
            try:
                from openpyxl.utils.cell import range_boundaries
                min_col, min_row, max_col, max_row = range_boundaries(merged_range_str)
            except ImportError:
                # 兼容旧版 openpyxl 的备用方案
                col_start, row_start, col_end, row_end = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", merged_range_str).groups()

                min_col = column_index_from_string(col_start)
                min_row = int(row_start)
                max_col = column_index_from_string(col_end)
                max_row = int(row_end)

            # 获取合并区域的左上角单元格
            top_left_cell = sheet.cell(row=min_row, column=min_col)
            top_left_cell_value = top_left_cell.value
            top_left_cell_format = top_left_cell.number_format # 获取源单元格的数字格式

            # 取消合并单元格
            sheet.unmerge_cells(merged_range_str)

            # 将值向下填充，并将数字格式复制到所有先前合并的单元格
            for row_idx in range(min_row, max_row + 1):
                for col_idx in range(min_col, max_col + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.value = top_left_cell_value
                    cell.number_format = top_left_cell_format # 将源单元格的格式复制过来

        # 保存修改后的工作簿
        workbook.save(output_filepath)
        print(f"成功处理并保存到: {output_filepath}")

    except FileNotFoundError:
        print(f"错误: 未找到输入文件 {input_filepath}")
    except Exception as e:
        print(f"处理文件 {input_filepath} 时发生错误: {e}")

def main():
    """
    主函数，用于对指定目录中所有 .xlsx 文件执行取消合并和填充脚本。
    """
    # 加载 .env 文件中的环境变量
    load_dotenv() 

    # 从环境变量获取输入和输出目录
    # 如果环境变量未设置，可以使用 .get() 方法提供默认值
    input_directory = os.getenv("INPUT_DIRECTORY") 
    output_directory = os.getenv("OUTPUT_DIRECTORY")

    print(input_directory, output_directory)

    # 如果输出目录不存在则创建它
    os.makedirs(output_directory, exist_ok=True)

    # 遍历输入目录中的所有文件
    for filename in os.listdir(input_directory):
        # 检查文件是否是 .xlsx 文件
        if filename.endswith(".xlsx"):
            input_file_path = os.path.join(input_directory, filename)
            output_file_path = os.path.join(output_directory, filename)
            
            print(f"正在处理文件: {filename}")
            unmerge_and_fill_with_original_format(input_file_path, output_file_path)

if __name__ == "__main__":
    main()