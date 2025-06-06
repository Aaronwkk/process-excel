import pandas as pd
from pymongo import MongoClient
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.merge import MergeCells
from openpyxl.utils import range_boundaries
from dotenv import load_dotenv
import os

load_dotenv() # 这会加载 .env 文件中的所有变量到 os.environ

def apply_styles(ws):
    """
    应用 Excel 文件的样式设置。
    Args:
        ws: openpyxl 的 worksheet 对象。
    """
    # 获取工作表的实际最大列数
    max_column = ws.max_column
    max_column_letter = get_column_letter(max_column)

    # --- 前三行合并单元格，字体大小为24，加粗 ---
    # 合并 A1 到 A3，横跨所有列
    ws.merge_cells(f'A1:{max_column_letter}3')
    top_left_cell = ws['A1']
    top_left_cell.font = Font(size=24, bold=True)
    top_left_cell.alignment = Alignment(horizontal='center', vertical='center') # 居中

    # --- 第四行合并单元格 ---
    ws.merge_cells(f'A4:{max_column_letter}4')
    ws['A4'].alignment = Alignment(horizontal='center', vertical='center') # 居中

    # --- 第五行和第六行纵向合并单元格，作为表格表头，字体大小12 并加粗 ---
    # 遍历所有列，对第五行和第六行进行纵向合并
    for col_idx in range(1, max_column + 1):
        col_letter = get_column_letter(col_idx)
        # 检查是否已经存在合并单元格，避免重复合并
        # 这里需要更精细地检查，避免影响到A1:MaxCol3 和 A4:MaxCol4 的合并
        # 我们可以检查当前单元格是否属于这些大合并单元格的一部分
        is_part_of_large_merge = False
        for merged_range in ws.merged_cells.ranges:
            min_col, min_row, max_col_m, max_row_m = range_boundaries(str(merged_range))
            if min_row <= 5 <= max_row_m and min_col <= col_idx <= max_col_m:
                is_part_of_large_merge = True
                break
        
        if not is_part_of_large_merge: # 只有当不属于大合并单元格时才进行纵向合并
            if f'{col_letter}5:{col_letter}6' not in [str(m) for m in ws.merged_cells.ranges]:
                ws.merge_cells(f'{col_letter}5:{col_letter}6')
            
            # 设置表头字体和居中
            header_cell = ws.cell(row=5, column=col_idx)
            header_cell.font = Font(size=12, bold=True)
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 确保第六行的对应单元格也居中，因为它们被合并了
            ws.cell(row=6, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
        else:
            # 如果是大的合并单元格的一部分，仍然确保内容居中（尽管通常已经设置过了）
            cell_5 = ws.cell(row=5, column=col_idx)
            cell_5.alignment = Alignment(horizontal='center', vertical='center')
            cell_6 = ws.cell(row=6, column=col_idx)
            cell_6.alignment = Alignment(horizontal='center', vertical='center')


    # --- 设置所有单元格文字居中 ---
    # 遍历所有行和列，但要避免修改已经合并的标题单元格的对齐方式，
    # 因为它们的对齐方式已经在上面显式设置过了。
    # 我们可以从第 5 行开始遍历数据行，确保数据行的对齐方式
    for r_idx in range(5, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=r_idx, column=col_idx)
            # 避免覆盖顶部大标题的对齐方式，只设置未合并单元格或数据单元格
            is_merged_cell_part = False
            for merged_range in ws.merged_cells.ranges:
                min_col, min_row, max_col_m, max_row_m = range_boundaries(str(merged_range))
                if min_row <= r_idx <= max_row_m and min_col <= col_idx <= max_col_m and (min_row < 5 or max_row_m < 6): # 排除第五行和第六行表头合并
                    is_merged_cell_part = True
                    break
            if not is_merged_cell_part:
                cell.alignment = Alignment(horizontal='center', vertical='center')


    # --- 列宽设定 ---
    # 对于所有列，设置默认宽度为12
    for col_idx in range(1, max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    # “身份证号码”列宽为 20
    # 需要找到“身份证号码”列的索引
    header_row_index = 5
    # 重新从工作表中获取表头，因为可能新增了列
    headers_in_excel = [cell.value for cell in ws[header_row_index]]
    
    # 清理表头，去除空格、换行符，以便查找
    cleaned_headers_in_excel = [str(h).strip().replace('\n', '').replace('\r', '') if h is not None else '' for h in headers_in_excel]

    try:
        id_card_col_idx = cleaned_headers_in_excel.index("身份证号码") + 1
        ws.column_dimensions[get_column_letter(id_card_col_idx)].width = 20
    except ValueError:
        print("警告: 未找到 '身份证号码' 列，无法单独设置其宽度。")


def main():
    # --- 配置参数 ---
    
    mongodb_uri = os.getenv("MONGODB_URI")
    db_name = os.getenv("DB_NAME")
    collection_name = os.getenv("COLLECTION_NAME")

    path = os.getenv("DATA_DIRECTORY") # 存放Excel文件的目录
    output_path = os.getenv("_DATA_DIRECTORY") # 存放Excel文件的目录
    INSURANCE_AMOUNT_FACTOR = int(os.environ.get("INSURANCE_AMOUNT_FACTOR", "17"))

    # 确保输出目录存在
    os.makedirs(output_path, exist_ok=True)

    # 连接 MongoDB
    client = MongoClient(mongodb_uri)
    db = client[db_name]
    collection = db[collection_name]

    # 遍历处理目录下的所有 Excel 文件
    for filename in os.listdir(path):
        if filename.endswith(('.xlsx', '.xls')):
            file_path = os.path.join(path, filename)
            print(f"正在处理文件: {filename}")

            try:
                # 提取行政村关键字
                match = re.match(r"^(.*?村)", filename)
                if not match:
                    print(f"警告: 文件名 '{filename}' 未能提取到行政村信息，跳过。")
                    continue
                village_name = match.group(1).replace("村委会", "村")

                # 从 MongoDB 查找数据
                mongo_data = list(collection.find({"village": village_name}))
                if not mongo_data:
                    print(f"警告: 在 MongoDB 中未找到与 '{village_name}' 匹配的数据，跳过文件 '{filename}'。")
                    continue

                # 加载 Excel 文件 (使用 openpyxl 进行写入和格式化)
                wb = load_workbook(file_path)
                ws = wb.active

                # 找到表头行（第五行）
                header_row_index = 5
                # 获取原始的表头，用于确定现有列的数量
                original_headers = [cell.value for cell in ws[header_row_index]]
                
                # 清理原始表头，去除空格、换行符
                cleaned_original_headers = [str(h).strip().replace('\n', '').replace('\r', '') if h is not None else '' for h in original_headers]

                # --- 新增“赔款金额”列 ---
                new_col_name_payment = "赔款金额"
                payment_amount_col_idx = -1 # 初始化，表示未找到
                
                # 检查“赔款金额”是否已存在
                if new_col_name_payment not in cleaned_original_headers:
                    # 如果不存在，则在现有列的末尾添加新列
                    payment_amount_col_idx = len(original_headers) + 1
                    ws.cell(row=header_row_index, column=payment_amount_col_idx, value=new_col_name_payment)
                    # 同时更新第六行，因为第五行和第六行会合并
                    ws.cell(row=header_row_index + 1, column=payment_amount_col_idx, value=new_col_name_payment)
                else:
                    # 如果已存在，找到其索引
                    payment_amount_col_idx = cleaned_original_headers.index(new_col_name_payment) + 1

                # --- 查找或新增“损失程度”列 ---
                loss_degree_col_name = "损失程度"
                loss_degree_col_idx = -1 # 初始化
                
                # 重新获取当前最新的表头（可能已经添加了“赔款金额”）
                current_headers_for_loss_degree_check = [cell.value for cell in ws[header_row_index]]
                cleaned_current_headers_for_loss_degree_check = [str(h).strip().replace('\n', '').replace('\r', '') if h is not None else '' for h in current_headers_for_loss_degree_check]

                if loss_degree_col_name not in cleaned_current_headers_for_loss_degree_check:
                    # 如果不存在，则在当前列的末尾添加新列
                    loss_degree_col_idx = len(current_headers_for_loss_degree_check) + 1
                    ws.cell(row=header_row_index, column=loss_degree_col_idx, value=loss_degree_col_name)
                    # 同时更新第六行
                    ws.cell(row=header_row_index + 1, column=loss_degree_col_idx, value=loss_degree_col_name)
                else:
                    # 如果已存在，找到其索引
                    loss_degree_col_idx = cleaned_current_headers_for_loss_degree_check.index(loss_degree_col_name) + 1


                # 现在，重新获取完整的、最新的表头，用于查找“被保险人”和“投保面积”的索引
                # 确保在所有新列添加完毕后再获取一次，这样索引才是正确的
                final_headers = [cell.value for cell in ws[header_row_index]]
                cleaned_final_headers = [str(h).strip().replace('\n', '').replace('\r', '') if h is not None else '' for h in final_headers]


                # 查找相关列的索引
                try:
                    insured_person_col_idx = cleaned_final_headers.index("被保险人") + 1 # +1 是因为 openpyxl 是从 1 开始计数
                    insurance_area_col_idx = cleaned_final_headers.index("投保面积") + 1
                except ValueError as e:
                    print(f"错误: 文件 '{filename}' 中缺少必要的列 '被保险人' 或 '投保面积'。{e}")
                    continue
                
                # 定义浅黄色填充
                light_yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

                # 遍历数据行（从第六行开始）
                # 注意：数据从 header_row_index + 1 开始，即第 6 行
                for r_idx in range(header_row_index + 1, ws.max_row + 1):
                    # 检查是否是空行
                    if all(cell.value is None for cell in ws[r_idx]):
                        continue

                    # 获取当前行的“被保险人”和“投保面积”
                    insured_person = ws.cell(row=r_idx, column=insured_person_col_idx).value
                    insurance_area = ws.cell(row=r_idx, column=insurance_area_col_idx).value

                    # 处理“被保险人”字段的潜在类型问题
                    if insured_person is not None:
                        insured_person = str(insured_person).strip()

                    # 填充“赔款金额”
                    # 确保只填充到“赔款金额”列
                    if isinstance(insurance_area, (int, float)):
                        ws.cell(row=r_idx, column=payment_amount_col_idx, value=insurance_area * INSURANCE_AMOUNT_FACTOR)
                    else:
                        ws.cell(row=r_idx, column=payment_amount_col_idx, value="") # 如果投保面积无效，则留空

                    # 填充“损失程度”
                    loss_percentage_value = ""
                    found_match = False
                    for data_item in mongo_data:
                        # 处理MongoDB中可能存在的类型问题
                        farmer_name_from_db = str(data_item.get("farmer_name", "")).strip()
                        if insured_person == farmer_name_from_db:
                            loss_percentage = data_item.get("loss_percentage")
                            if isinstance(loss_percentage, (int, float)):
                                # 四舍五入到小数点后一位，然后乘以100格式化为百分比
                                loss_percentage_value = f"{round(loss_percentage * 100, 1):.1f}%"
                            else:
                                loss_percentage_value = ""
                            ws.cell(row=r_idx, column=loss_degree_col_idx, value=loss_percentage_value)
                            # 设置背景色为浅黄色
                            for cell in ws[r_idx]:
                                cell.fill = light_yellow_fill
                            found_match = True
                            break

                    if not found_match and mongo_data:
                        # 如果没有匹配，填写任意一个 avg_loss_same_level
                        avg_loss_same_level = mongo_data[0].get("avg_loss_same_level")
                        if isinstance(avg_loss_same_level, (int, float)):
                            # 四舍五入到小数点后一位，然后乘以100格式化为百分比
                            loss_percentage_value = f"{round(avg_loss_same_level * 100, 1):.1f}%"
                        else:
                            loss_percentage_value = ""
                        ws.cell(row=r_idx, column=loss_degree_col_idx, value=loss_percentage_value)
                    elif not found_match and not mongo_data:
                        ws.cell(row=r_idx, column=loss_degree_col_idx, value="") # 如果MongoDB数据为空，则留空

                # 应用样式
                apply_styles(ws)

                # 保存处理后的文件
                output_file_path = os.path.join(output_path, filename)
                wb.save(output_file_path)
                print(f"文件 '{filename}' 处理完成，已保存到: {output_file_path}")

            except Exception as e:
                print(f"处理文件 '{filename}' 时发生错误: {e}")

    client.close()
    print("所有文件处理完毕。")

if __name__ == "__main__":
    main()